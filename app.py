from flask import Flask, request, jsonify, send_file
from flask_cors import CORS
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime, timedelta
import io

app = Flask(__name__)
CORS(app)

# ── COLORS ───────────────────────────────────────────────────
FILLS = {
    'inp':     PatternFill('solid', start_color='FEF9C3'),  # yellow
    'nurs':    PatternFill('solid', start_color='DCFCE7'),  # green
    'pc':      PatternFill('solid', start_color='D1FAE5'),  # mint
    'ed':      PatternFill('solid', start_color='DBEAFE'),  # blue
    'sub':     PatternFill('solid', start_color='EDE9FE'),  # purple
    'math':    PatternFill('solid', start_color='FCE7F3'),  # pink
    'hol':     PatternFill('solid', start_color='F3F4F6'),  # gray
    'or':      PatternFill('solid', start_color='F1F5F9'),  # light
    'shelf':   PatternFill('solid', start_color='FEE2E2'),  # red
    'read':    PatternFill('solid', start_color='FFF7ED'),  # orange
    'hdr':     PatternFill('solid', start_color='1A1A1A'),
    'col_hdr': PatternFill('solid', start_color='E8E8E5'),
    'sec_hdr': PatternFill('solid', start_color='374151'),
    'wknd':    PatternFill('solid', start_color='F0FDF4'),
}
FONT_COLORS = {
    'inp':'854D0E','nurs':'15803D','pc':'065F46','ed':'1E40AF',
    'sub':'5B21B6','math':'9D174D','hol':'6B7280','or':'475569',
    'shelf':'991B1B','read':'9A3412','hdr':'FFFFFF','col_hdr':'374151',
    'sec_hdr':'FFFFFF','wknd':'166534',
}

def rkey(v):
    if not v: return 'or'
    if 'Inpatient' in v: return 'inp'
    if 'Nursery' in v:   return 'nurs'
    if ' PC' in v:       return 'pc'
    if ' ED' in v:       return 'ed'
    if 'Sub' in v:       return 'sub'
    if 'Matheny' in v:   return 'math'
    if 'Holiday' in v:   return 'hol'
    if 'Shelf' in v:     return 'shelf'
    if 'Reading' in v:   return 'read'
    return 'or'

thin = Side(style='thin', color='CCCCCC')
med  = Side(style='medium', color='999999')
def border(): return Border(left=thin,right=thin,top=thin,bottom=thin)
def med_border(): return Border(left=med,right=med,top=med,bottom=med)

def sc(ws, row, col, value, key='or', bold=False, size=10, halign='center', wrap=False):
    c = ws.cell(row=row, column=col, value=value)
    c.fill = FILLS.get(key, FILLS['or'])
    c.font = Font(name='Arial', size=size, bold=bold, color=FONT_COLORS.get(key,'1A1A1A'))
    c.alignment = Alignment(horizontal=halign, vertical='center', wrap_text=wrap)
    c.border = border()
    return c

def pc(ws, row, col, value, bold=False, size=10, halign='left'):
    c = ws.cell(row=row, column=col, value=value)
    c.font = Font(name='Arial', size=size, bold=bold)
    c.alignment = Alignment(horizontal=halign, vertical='center')
    return c

def get_date(iso, wi, di):
    d = datetime.strptime(iso, '%Y-%m-%d') + timedelta(weeks=wi, days=di)
    return d

def week_dates(iso, wi):
    d = get_date(iso, wi, 0)
    e = get_date(iso, wi, 4)
    return f"{d.month}/{d.day}–{e.month}/{e.day}"

def wed_date(iso, wi):
    d = get_date(iso, wi, 2)
    return f"{d.month}/{d.day}"

def day_iso(iso, wi, di):
    return get_date(iso, wi, di).strftime('%Y-%m-%d')

GROUP_LABELS = {
    'nbimc_all':'NBIMC (all 6 wks)',
    'nbimc_uh_ed':'NBIMC / UH ED',
    'nbimc_cbmc_sub':'NBIMC / CBMC Sub',
    'uh_cbmc_inp':'UH / CBMC Inpatient',
    'uh_cbmc_both':'UH / CBMC Inp + Sub',
    'overflow':'Overflow',
}

TIMING = {
    'NBIMC Inpatient':'6:30am sign-out–5pm weekdays; 6:30am–6:30pm short-call; weekends 8am–4pm',
    'CBMC Inpatient':'6:30am–4pm weekdays; 6:30am–6:30pm short-call; weekends 8am–4pm',
    'NBIMC PC':'8:45am (huddle)',
    'UH PC':'DOC 5100 · 9am–5pm',
    'NBIMC ED':'7am–3pm or 3pm–11pm (Fridays 12–6pm)',
    'UH ED':'C-level behind main ED · 7am–3pm or 3pm–11pm (Fridays 12–6pm)',
    'NBIMC Nursery':'7am–5pm',
    'UH Nursery':'F-level Orange · 7am–4pm',
    'NBIMC Subspecialty':'9am',
    'UH Subspecialty':'DOC 4300 · 9am–5pm',
    'CBMC Subspecialty':'See CBMC Schedule (contact Karen Feniello)',
}

def build_excel(data):
    students   = data['students']
    start_iso  = data['startISO']
    cohort     = data['cohort']
    shelf_date = data.get('shelfDate','')
    math_weeks = data['mathWeeks']  # 0-indexed
    holiday_dates = set(data.get('holidayDates', []))

    wb = Workbook()
    wb.remove(wb.active)

    # ── 1. STUDENT LIST ──────────────────────────────────────
    ws = wb.create_sheet('Student List')
    ws.column_dimensions['A'].width = 6
    ws.column_dimensions['B'].width = 28
    ws.column_dimensions['C'].width = 32
    for col,h in enumerate(['#','Name','Group'],1):
        sc(ws,1,col,h,'sec_hdr',bold=True,size=11)
    for i,s in enumerate(students,2):
        pc(ws,i,1,s['num'],halign='center')
        pc(ws,i,2,s['name'])
        pc(ws,i,3,s['groupLabel'])
        for col in range(1,4): ws.cell(row=i,column=col).border=border()
    ws.freeze_panes='A2'

    # ── 2. OVERVIEW ──────────────────────────────────────────
    ws = wb.create_sheet('Overview')
    ws.column_dimensions['A'].width=5
    ws.column_dimensions['B'].width=26
    ws.column_dimensions['C'].width=8
    for c in range(4,10): ws.column_dimensions[get_column_letter(c)].width=22
    for col,h in enumerate(['#','Name','Slot','Wk 1','Wk 2','Wk 3','Wk 4','Wk 5','Wk 6'],1):
        sc(ws,1,col,h,'sec_hdr',bold=True,size=10)
    for i,s in enumerate(students,2):
        pc(ws,i,1,s['num'],halign='center')
        pc(ws,i,2,s['name'])
        pc(ws,i,3,s['slotNum'],halign='center')
        for wi,wk in enumerate(s['weeks']):
            primary=next((c for c in wk if c and c!='Orientation' and 'Shelf' not in c and 'Reading' not in c and c!='Matheny' and c!='Holiday' and c!=''),'')
            primary=primary.replace(' (Call)','')
            has_math='Matheny' in wk
            label=primary+(' + Matheny' if has_math else '')
            sc(ws,i,4+wi,label,rkey(primary),size=9)
    ws.freeze_panes='D2'

    # ── 3. INDIVIDUAL SCHEDULE SHEETS by group ───────────────
    DAY_HDRS=['Dates','Monday','Tuesday','Wednesday','Thursday','Friday','Weekend Call']
    group_order=['nbimc_all','nbimc_uh_ed','nbimc_cbmc_sub','uh_cbmc_inp','uh_cbmc_both','overflow']
    from collections import defaultdict
    by_group=defaultdict(list)
    for s in students: by_group[s['group']].append(s)

    for gid in group_order:
        grp=by_group.get(gid,[])
        if not grp: continue
        sheet_name=(GROUP_LABELS.get(gid,gid)).replace('/','-').replace('  ',' ')[:31]
        ws=wb.create_sheet(sheet_name)
        ws.column_dimensions['A'].width=12
        for c in range(2,8): ws.column_dimensions[get_column_letter(c)].width=20

        row=1
        for s in grp:
            # Name header
            c=ws.cell(row=row,column=1,value=f"{s['num']}. {s['name']}  |  Slot {s['slotNum']}  |  {cohort}")
            c.fill=FILLS['hdr']; c.font=Font(name='Arial',size=11,bold=True,color='FFFFFF')
            c.alignment=Alignment(horizontal='left',vertical='center')
            ws.merge_cells(start_row=row,start_column=1,end_row=row,end_column=7)
            ws.row_dimensions[row].height=18; row+=1

            # Group label
            c2=ws.cell(row=row,column=1,value=s['groupLabel'])
            c2.fill=FILLS['col_hdr']; c2.font=Font(name='Arial',size=9,color='6B7280')
            c2.alignment=Alignment(horizontal='left',vertical='center')
            ws.merge_cells(start_row=row,start_column=1,end_row=row,end_column=7)
            row+=1

            # Column headers
            for col,h in enumerate(DAY_HDRS,1):
                sc(ws,row,col,h,'col_hdr',bold=True,size=9)
            row+=1

            # 6 week rows
            for wi,wk in enumerate(s['weeks']):
                dates=week_dates(start_iso,wi)
                sc(ws,row,1,dates,'col_hdr',size=9,halign='left')
                for di in range(5):
                    content=wk[di] if wk[di] else ''
                    if di>0 and day_iso(start_iso,wi,di) in holiday_dates:
                        content='Holiday'
                    k=rkey(content)
                    if content=='Holiday': k='hol'
                    sc(ws,row,2+di,content,k,size=9,wrap=True)
                wknd=wk[5] if len(wk)>5 else ''
                sc(ws,row,7,wknd,'wknd' if wknd else 'or',size=9)
                ws.row_dimensions[row].height=28; row+=1

            # Timing sidebar (col 9+)
            seen=set()
            for wk in s['weeks']:
                for cell in wk:
                    if cell:
                        base=cell.replace(' (Call)','').replace(' (7am-3pm)','').replace(' (3pm-11pm)','').replace(' (12pm-6pm)','')
                        seen.add(base)
            timing_start=row-6
            tr=timing_start
            for base in [b for b in seen if b in TIMING]:
                lc=ws.cell(row=tr,column=9,value=base)
                lc.font=Font(name='Arial',size=9,bold=True)
                vc=ws.cell(row=tr,column=10,value=TIMING[base])
                vc.font=Font(name='Arial',size=9,color='555555')
                tr+=1
            ws.column_dimensions['I'].width=22
            ws.column_dimensions['J'].width=40

            row+=1  # blank between students

    # ── 4. MATHENY SCHEDULE ──────────────────────────────────
    ws=wb.create_sheet('Matheny Schedule')
    ws.column_dimensions['A'].width=26
    ws.column_dimensions['B'].width=26
    ws.column_dimensions['C'].width=26
    sc(ws,1,1,'Matheny Schedule','sec_hdr',bold=True,size=12)
    ws.merge_cells('A1:C1')
    for gi,mw in enumerate(math_weeks):
        wd=wed_date(start_iso,mw)
        sc(ws,2,gi+1,f"Week {mw+1} — Wed {wd}",'col_hdr',bold=True,size=10)
    ws.row_dimensions[2].height=20
    groups_math=[[],[],[]]
    for s in students:
        gi=math_weeks.index(s['mathenyWeek']) if s['mathenyWeek'] in math_weeks else 0
        groups_math[gi].append(s['name'])
    max_len=max(len(g) for g in groups_math) if any(groups_math) else 0
    for ri in range(max_len):
        for gi in range(3):
            name=groups_math[gi][ri] if ri<len(groups_math[gi]) else ''
            c=ws.cell(row=3+ri,column=gi+1,value=name)
            c.font=Font(name='Arial',size=10)
            c.border=border()
            c.alignment=Alignment(horizontal='left',vertical='center')

    # ── 5. SECTION DATES ─────────────────────────────────────
    SITE_ROTS={
        'NBIMC':['NBIMC Inpatient','NBIMC Nursery','NBIMC PC','NBIMC ED','NBIMC Subspecialty'],
        'CBMC':['CBMC Inpatient','CBMC Subspecialty'],
        'UH':['UH Nursery','UH PC','UH ED','UH Subspecialty'],
    }
    for site,rots in SITE_ROTS.items():
        ws=wb.create_sheet(f'{site} Section Dates')
        ws.column_dimensions['A'].width=10
        ws.column_dimensions['B'].width=14
        for c in range(3,15): ws.column_dimensions[get_column_letter(c)].width=20
        cur=1

        for rot in rots:
            is_inp='Inpatient' in rot
            is_ed=' ED' in rot
            k=rkey(rot)

            # Section header
            c=ws.cell(row=cur,column=1,value=rot)
            c.fill=FILLS['sec_hdr']; c.font=Font(name='Arial',size=11,bold=True,color='FFFFFF')
            c.alignment=Alignment(horizontal='left',vertical='center')
            ws.merge_cells(start_row=cur,start_column=1,end_row=cur,end_column=10)
            cur+=1

            if is_inp:
                # Header
                sc(ws,cur,1,'Weeks','col_hdr',bold=True,size=9,halign='left')
                sc(ws,cur,2,'Dates','col_hdr',bold=True,size=9)
                for ci in range(6): sc(ws,cur,3+ci,rot,'col_hdr',bold=True,size=9)
                cur+=1
                for w1,w2 in [(0,1),(2,3),(4,5)]:
                    d1=get_date(start_iso,w1,0); d2=get_date(start_iso,w2,4)
                    dlabel=f"{d1.month}/{d1.day}–{d2.month}/{d2.day}"
                    in_pair=[s for s in students if any(c2 for wk in [s['weeks'][w1],s['weeks'][w2]] for c2 in wk if rot in c2)]
                    sc(ws,cur,1,f"{w1+1} & {w2+1}",'col_hdr',bold=True,size=9,halign='left')
                    sc(ws,cur,2,dlabel,k,size=9)
                    for ci,s in enumerate(in_pair):
                        sc(ws,cur,3+ci,s['name'],k,size=9,halign='left')
                    cur+=1
            else:
                if is_ed:
                    sc(ws,cur,1,'Week','col_hdr',bold=True,size=9,halign='left')
                    sc(ws,cur,2,'Dates','col_hdr',bold=True,size=9)
                    sc(ws,cur,3,rot+' 7am-3pm','col_hdr',bold=True,size=9)
                    sc(ws,cur,4,rot+' 7am-3pm','col_hdr',bold=True,size=9)
                    sc(ws,cur,5,rot+' 3pm-11pm','col_hdr',bold=True,size=9)
                    sc(ws,cur,6,rot+' 3pm-11pm','col_hdr',bold=True,size=9)
                    cur+=1
                else:
                    max_n=max((len([s for s in students if any(rot in c2 for c2 in s['weeks'][wi])]) for wi in range(6)),default=1)
                    sc(ws,cur,1,'Week','col_hdr',bold=True,size=9,halign='left')
                    sc(ws,cur,2,'Dates','col_hdr',bold=True,size=9)
                    for ci in range(max(max_n,1)): sc(ws,cur,3+ci,rot,'col_hdr',bold=True,size=9)
                    cur+=1

                for wi in range(6):
                    in_wk=[s for s in students if any(rot in c2 for c2 in s['weeks'][wi])]
                    if not in_wk: continue
                    dates=week_dates(start_iso,wi)
                    sc(ws,cur,1,f"Wk {wi+1}",'col_hdr',bold=True,size=9,halign='left')
                    sc(ws,cur,2,dates,k,size=9)
                    if is_ed:
                        am=[s for s in in_wk if any(rot in c2 and '7am' in c2 for c2 in s['weeks'][wi])]
                        pm=[s for s in in_wk if any(rot in c2 and '3pm-11pm' in c2 for c2 in s['weeks'][wi])]
                        for ci,s in enumerate([*am,*pm]): sc(ws,cur,3+ci,s['name'],k,size=9,halign='left')
                    else:
                        for ci,s in enumerate(in_wk): sc(ws,cur,3+ci,s['name'],k,size=9,halign='left')
                    cur+=1

            # Timing note
            if rot in TIMING:
                c=ws.cell(row=cur,column=1,value=f"{rot}: {TIMING[rot]}")
                c.font=Font(name='Arial',size=8,italic=True,color='6B7280')
                ws.merge_cells(start_row=cur,start_column=1,end_row=cur,end_column=10)
            cur+=2  # blank row between sections

    buf=io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf

@app.route('/generate', methods=['POST'])
def generate():
    try:
        data=request.get_json()
        buf=build_excel(data)
        cohort=data.get('cohort','schedule').replace(' ','_').replace('—','').strip('_')
        filename=f"peds_schedule_{cohort}.xlsx"
        return send_file(buf, as_attachment=True, download_name=filename,
                        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    except Exception as e:
        return jsonify({'error':str(e)}), 500

@app.route('/health', methods=['GET'])
def health():
    return jsonify({'status':'ok'})

if __name__=='__main__':
    app.run(debug=False, host='0.0.0.0', port=5000)
